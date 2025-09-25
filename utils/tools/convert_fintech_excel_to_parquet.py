from __future__ import annotations
from pathlib import Path
import re
import pandas as pd
from openpyxl import load_workbook

# Exact filename per your screenshot
EXCEL = Path("data") / "Fintech Search V15.xlsm"
OUT_DIR = Path("data")
OUT_DIR.mkdir(exist_ok=True)

# ---- robust sheet finders ----
def find_sheet_name(names, *candidates) -> str | None:
    """
    Try exact, then case-insensitive, then fuzzy contains for any candidate tokens.
    Example: find_sheet_name(all_names, "Fintech ID Cards")
    Example: find_sheet_name(all_names, ("Data","Structure","2"))
    """
    # exact
    for n in names:
        if n in candidates:
            return n
    # case-insensitive exact
    low_map = {n.lower(): n for n in names}
    for c in candidates:
        if isinstance(c, str) and c.lower() in low_map:
            return low_map[c.lower()]
    # fuzzy: require all tokens to appear (order agnostic)
    toks_list: list[list[str]] = []
    for c in candidates:
        if isinstance(c, str):
            toks_list.append([c.lower()])
        elif isinstance(c, (tuple, list)):
            toks_list.append([t.lower() for t in c])
    for n in names:
        nlow = n.lower()
        for toks in toks_list:
            if all(t in nlow for t in toks):
                return n
    return None

def convert():
    if not EXCEL.exists():
        raise FileNotFoundError(f"Excel not found at {EXCEL} (expected per your layout).")

    # open once for sheet names + hyperlinks
    wb = load_workbook(EXCEL, read_only=False, keep_links=True)
    all_names = [s.title for s in wb.worksheets]  # preserves case as seen in Excel

    sheet_cards = find_sheet_name(
        all_names,
        "Fintech ID Cards",           # expected
        ("fintech", "id", "card"),    # fuzzy fallback
    )
    sheet_ds2 = find_sheet_name(
        all_names,
        "Data Structure 2",           # expected
        ("data", "structure", "2"),   # fuzzy fallback
        ("data", "structure", "ii"),  # just in case roman numeral
    )

    if sheet_cards is None or sheet_ds2 is None:
        print("Could not find sheets with the expected names.")
        print("Workbook sheet names I can see:")
        for n in all_names:
            print(" -", n)
        raise SystemExit(1)

    ws_cards = wb[sheet_cards]
    ws_ds2   = wb[sheet_ds2]

    # -------- 1) Companies table (details) --------
    rows = []
    max_row_cards = ws_cards.max_row
    for r in range(4, max_row_cards + 1):   # data typically starts at row 4
        name = ws_cards.cell(r, 1).value    # Col A: company name
        if name is None or str(name).strip() == "":
            break
        creation  = ws_cards.cell(r, 2).value
        employees = ws_cards.cell(r, 3).value
        funding   = ws_cards.cell(r, 4).value
        country   = ws_cards.cell(r, 5).value
        notes     = ws_cards.cell(r, 6).value
        desc      = ws_cards.cell(r, 7).value

        rows.append({
            "Name": str(name).strip(),
            "Creation": creation,
            "Employees": employees,
            "Funding ($m)": funding,
            "Country": (country or ""),
            "Notes": (notes or ""),
            "Description": str(desc or "").replace("\n", " ").replace("\r", " "),
        })
    companies = pd.DataFrame(rows)

    # -------- 2) Categories table (Team/Function/Subfunction + Name) --------
    # header=2 matches your current app; adjust if your header row changes.
    df_ds2 = pd.read_excel(EXCEL, sheet_name=sheet_ds2, header=2, engine="openpyxl")
    df_ds2.columns = [str(c).strip() for c in df_ds2.columns]

    # Normalize known columns
    col_map = {"Sector": "Team", "sub-sector": "Function", "granularity": "Subfunction", "FINTECH": "Name"}
    df_ds2 = df_ds2.rename(columns={k: v for k, v in col_map.items() if k in df_ds2.columns})

    keep_cols = [c for c in ["Team", "Function", "Subfunction", "Name"] if c in df_ds2.columns]
    if not keep_cols or "Name" not in keep_cols:
        raise ValueError(f"'{sheet_ds2}' sheet does not appear to contain Team/Function/Subfunction/FINTECH.")

    df_ds2 = df_ds2[keep_cols].copy()
    df_ds2 = df_ds2[df_ds2["Name"].notna() & (df_ds2["Name"].astype(str).str.strip() != "")]
    for c in keep_cols:
        df_ds2[c] = df_ds2[c].astype(str).str.strip()

    # -------- website hyperlinks from DS2 Name column (assumed col 5) --------
    link_by_name = {}
    for r in range(4, ws_ds2.max_row + 1):
        cell = ws_ds2.cell(r, 5)  # Column E in your layout (FINTECH/Name)
        nm = cell.value
        if nm is None or str(nm).strip() == "":
            break
        nm = str(nm).strip()
        url = ""
        if cell.hyperlink:
            url = cell.hyperlink.target or ""
        else:
            # handle =HYPERLINK("url","text")
            v = cell.value
            if isinstance(v, str) and v.upper().startswith("=HYPERLINK("):
                try:
                    url = v.split('"')[1]
                except Exception:
                    url = ""
        if nm not in link_by_name:
            link_by_name[nm] = url

    companies["Website"] = companies["Name"].map(link_by_name).fillna("")

    # numeric typing for sortability
    for col in ["Creation", "Employees", "Funding ($m)"]:
        if col in companies.columns:
            companies[col] = pd.to_numeric(companies[col], errors="coerce")

    # -------- write parquet --------
    (OUT_DIR / "companies.parquet").unlink(missing_ok=True)
    (OUT_DIR / "categories.parquet").unlink(missing_ok=True)

    companies.to_parquet(OUT_DIR / "companies.parquet", index=False)
    df_ds2.to_parquet(OUT_DIR / "categories.parquet", index=False)

    print(f"OK • Wrote {OUT_DIR/'companies.parquet'} ({len(companies)} rows)")
    print(f"OK • Wrote {OUT_DIR/'categories.parquet'} ({len(df_ds2)} rows)")
    print(f"Sheets used • Cards: '{sheet_cards}', DS2: '{sheet_ds2}'")

if __name__ == "__main__":
    convert()
