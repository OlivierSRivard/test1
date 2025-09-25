import streamlit as st
import pandas as pd
pd.set_option('display.max_rows', 300)       # or a big number like 1000
from pathlib import Path
from openpyxl import load_workbook

st.set_page_config(page_title="Company Search", page_icon="🔎", layout="wide")
st.title("Search Companies in Ecosystem")

# Define possible paths to the Excel data file
DATA_PATHS = [
    Path("data") / "Fintech Search V15.xlsm",
    Path("Fintech Search V15.xlsm"),
]

@st.cache_data(show_spinner=False)
def load_data():
    """Load data from the Excel file: Fintech details, category mappings, and company website links."""
    # Locate the Excel file
    excel_path = None
    for p in DATA_PATHS:
        if p.exists():
            excel_path = p
            break
    if not excel_path:
        raise FileNotFoundError("Could not find 'Fintech Search V15.xlsm'. Place it in 'data/' or project root.")
    
    # Load the workbook with openpyxl to access hyperlinks
    wb = load_workbook(excel_path, read_only=False, keep_links=True)
    ws_cards = wb["Fintech ID Cards"]
    ws_ds2 = wb["Data Structure 2"]
    
    # 1. Extract firm details from "Fintech ID Cards" sheet
    details_by_name = {}
    # Find last row in Fintech ID Cards by scanning column A (Name)
    max_row_cards = ws_cards.max_row
    # It's safer to stop at the last non-empty Name; we'll break when we hit a blank.
    for row in range(4, max_row_cards + 1):  # data starts at row 4
        name = ws_cards.cell(row, 1).value
        if name is None or str(name).strip() == "":
            break  # stop at first blank name (no more companies)
        creation = ws_cards.cell(row, 2).value
        employees = ws_cards.cell(row, 3).value
        funding = ws_cards.cell(row, 4).value
        country = ws_cards.cell(row, 5).value
        notes = ws_cards.cell(row, 6).value
        description = ws_cards.cell(row, 7).value
        
        # Convert numeric values to clean strings (avoid .0 for whole numbers, handle None)
        def fmt_num(x):
            if x is None:
                return ""
            # If float and is an integer (e.g., 3.0), convert to int
            if isinstance(x, float) and hasattr(x, "is_integer") and x.is_integer():
                x = int(x)
            return str(x)
        creation_str = fmt_num(creation)
        employees_str = fmt_num(employees)
        funding_str = fmt_num(funding)
        country_str = country or ""      # country is usually a short code, keep as string
        notes_str = notes or ""          # notes may be None
        desc_str = (description or "").replace("\n", " ").replace("\r", " ")
        
        details_by_name[str(name).strip()] = {
            "Creation": creation_str,
            "Employees": employees_str,
            "Funding ($m)": funding_str,
            "Country": country_str,
            "Notes": notes_str,
            "Description": desc_str
        }
    
    # 2. Extract category mappings and hyperlinks from "Data Structure 2" sheet
    # Read the sheet into a DataFrame for easy filtering/grouping
    df_ds2 = pd.read_excel(excel_path, sheet_name="Data Structure 2", header=2, engine="openpyxl")
    df_ds2.columns = [str(c).strip() for c in df_ds2.columns]  # normalize column names
    # Rename relevant columns for clarity
    col_map = {
        "Sector": "Team",
        "sub-sector": "Function",
        "granularity": "Subfunction",
        "FINTECH": "Name"
    }
    df_ds2 = df_ds2.rename(columns={k: v for k, v in col_map.items() if k in df_ds2.columns})
    # Drop unused columns (keep only Team, Function, Subfunction, Name)
    keep_cols = ["Team", "Function", "Subfunction", "Name"]
    df_ds2 = df_ds2[[c for c in keep_cols if c in df_ds2.columns]].copy()
    # Drop any rows with no company name (if any)
    df_ds2 = df_ds2[df_ds2["Name"].notna() & (df_ds2["Name"].str.strip() != "")]
    
    # Build a dictionary of website links for each company (hyperlinks in the "Name" cell)
    link_by_name = {}
    max_row_ds2 = ws_ds2.max_row
    for row in range(4, max_row_ds2 + 1):
        name_cell = ws_ds2.cell(row, 5)  # Column 5 is "FINTECH"/Name in Data Structure 2
        comp_name = name_cell.value
        if comp_name is None or str(comp_name).strip() == "":
            break  # no more entries
        comp_name = str(comp_name).strip()
        if comp_name not in link_by_name:  # only record the first hyperlink (should be the same for all entries of a company)
            if name_cell.hyperlink:
                link_by_name[comp_name] = name_cell.hyperlink.target or ""
            else:
                # If the Name cell has a HYPERLINK formula instead of a proper hyperlink object
                link_formula = name_cell.value
                if isinstance(link_formula, str) and link_formula.upper().startswith("=HYPERLINK("):
                    # Extract URL between the first pair of quotes in formula: =HYPERLINK("URL","text")
                    try:
                        url_part = link_formula.split('"')[1]
                        link_by_name[comp_name] = url_part
                    except Exception:
                        link_by_name[comp_name] = ""
        # continue looping to ensure we skip duplicates properly or capture later rows if first rows had no hyperlink
    
    # 3. Pre-compute unique category options and mappings for dynamic filters
    # Unique list of teams, functions, and subfunctions for multiselect options
    teams_list = sorted(df_ds2["Team"].dropna().unique().tolist())
    functions_list = sorted(df_ds2["Function"].dropna().unique().tolist())
    subfunctions_list = sorted(df_ds2["Subfunction"].dropna().unique().tolist())
    # Mappings for parent-child relationships
    team_to_functions = {}
    for team in teams_list:
        funcs = df_ds2[df_ds2["Team"] == team]["Function"].dropna().unique().tolist()
        team_to_functions[team] = sorted(funcs)
    function_to_subs = {}
    for func in functions_list:
        subs = df_ds2[df_ds2["Function"] == func]["Subfunction"].dropna().unique().tolist()
        function_to_subs[func] = sorted(subs)
    
    return details_by_name, df_ds2, link_by_name, team_to_functions, function_to_subs, teams_list, functions_list, subfunctions_list

# Load data (cached to avoid re-reading on every interaction)
try:
    details_by_name, df_ds2, link_by_name, team_to_functions, function_to_subs, teams_list, functions_list, subfunctions_list = load_data()
except Exception as e:
    st.error(f"Failed to load data: {e}")
    st.stop()

# ---------------------------
# Helpers for interactive tables
# ---------------------------

def _logo_url_from_link(url: str) -> str:
    try:
        from urllib.parse import urlparse
        domain = urlparse(url).netloc or ""
        domain = domain.lstrip("www.")
        return f"https://logo.clearbit.com/{domain}" if domain else ""
    except Exception:
        return ""

def build_results_df(result_names, include_categories=False, categories_by_name=None):
    """Construct a typed DataFrame for display with st.dataframe."""
    rows = []
    for name in sorted(result_names):
        info = details_by_name.get(name, {})
        website = link_by_name.get(name, "")
        rows.append({
            "Logo": _logo_url_from_link(website) if website else "",
            "Name": name,
            "Website": website,
            "Creation": info.get("Creation", ""),
            "Employees": info.get("Employees", ""),
            "Funding ($m)": info.get("Funding ($m)", ""),
            "Country": info.get("Country", ""),
            "Notes": info.get("Notes", ""),
            "Description": info.get("Description", ""),
            **({
                "Teams": (categories_by_name.get(name, {}).get("teams", "") if categories_by_name else ""),
                "Functions": (categories_by_name.get(name, {}).get("functions", "") if categories_by_name else ""),
                "Subfunctions": (categories_by_name.get(name, {}).get("subfunctions", "") if categories_by_name else "")
            } if include_categories else {})
        })
    df = pd.DataFrame(rows)

    # Coerce numeric columns so sorting works correctly
    for col in ["Creation", "Employees", "Funding ($m)"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Normalize empty countries to None for nicer filtering
    if "Country" in df.columns:
        df["Country"] = df["Country"].replace({"": None})

    return df

def render_results_df(df: pd.DataFrame, *, show_categories: bool = False, min_visible_rows: int = 50):
    """Add a Country filter and render a sortable, interactive table.

    Shows up to `min_visible_rows` rows without forcing an inner scroll;
    the page will scroll instead. If there are fewer rows, it shows them all.
    """
    # Country filter (unique values from current result set)
    if "Country" in df.columns:
        countries = sorted([c for c in df["Country"].dropna().unique().tolist() if str(c).strip() != ""])
        if countries:
            selected_countries = st.multiselect("Filter by Country", options=countries, default=countries)
            df = df[df["Country"].isin(selected_countries)]

    # Column configuration: clickable website + logos as images
    column_config = {
        "Logo": st.column_config.ImageColumn("Logo", width="small"),
        "Website": st.column_config.LinkColumn("Website"),
        "Creation": st.column_config.NumberColumn("Creation", help="Year of creation"),
        "Employees": st.column_config.NumberColumn("Employees"),
        "Funding ($m)": st.column_config.NumberColumn("Funding ($m)"),
        "Notes": st.column_config.TextColumn("Notes"),
        "Description": st.column_config.TextColumn("Description"),
        "Name": st.column_config.TextColumn("Name"),
        "Country": st.column_config.TextColumn("Country"),
    }

    display_cols = ["Logo", "Name", "Website", "Creation", "Employees", "Funding ($m)", "Country", "Notes", "Description"]
    if show_categories:
        display_cols += ["Teams", "Functions", "Subfunctions"]
        column_config.update({
            "Teams": st.column_config.TextColumn("Teams"),
            "Functions": st.column_config.TextColumn("Functions"),
            "Subfunctions": st.column_config.TextColumn("Subfunctions"),
        })

    # ---- NEW: compute a tall-enough height so the page scrolls (not the grid) ----
    n_rows = len(df)
    rows_to_show = max(1, min(n_rows, min_visible_rows))  # <= 50 or fewer if fewer results
    ROW_PX = 37      # approximate row height in px (adjust if your logos make rows taller)
    HEADER_PX = 38   # header bar height
    PADDING_PX = 16  # a little extra
    height = HEADER_PX + ROW_PX * rows_to_show + PADDING_PX
    # ------------------------------------------------------------------------------

    st.dataframe(
        df[display_cols],
        hide_index=True,
        width='stretch',
        column_config=column_config,
        height=height,  # <- key line: makes the component tall; the page will scroll
    )

# --- UI for search mode selection ---
search_mode = st.radio("Search by", ["Company Name", "Functional Categories", "Keyword"], index=0)

if search_mode == "Company Name":
    # Company Name Search
    firm_query = st.text_input("Enter part of the company name (case-insensitive):").strip()
    if firm_query:
        # Filter company names containing the query substring (case-insensitive)
        query_lower = firm_query.lower()
        result_names = [name for name in details_by_name.keys() if query_lower in name.lower()]
        result_count = len(result_names)
        st.caption(f"Found {result_count} matching firm(s)")
        if result_count == 0:
            st.info("No companies found. Try a different name or spelling.")
        else:
            # Build and display results table (interactive)
            df_results = build_results_df(result_names, include_categories=False)
            render_results_df(df_results, show_categories=False)

elif search_mode == "Functional Categories":
    # Functional Blocks Search (Team/Function/Subfunction)
    st.write("**Filter by functional classification:**")
    # Multi-select for Team(s)
    selected_teams = st.multiselect("Team", options=teams_list, default=[])
    # Update Function options based on selected teams (if any)
    if selected_teams:
        valid_functions = set()
        for team in selected_teams:
            for func in team_to_functions.get(team, []):
                valid_functions.add(func)
        function_options = sorted(valid_functions)
    else:
        function_options = functions_list  # all functions if no team filter
    selected_functions = st.multiselect("Function", options=function_options, default=[])
    # Update Subfunction options based on selected functions or teams
    if selected_functions:
        valid_subs = set()
        for func in selected_functions:
            for sub in function_to_subs.get(func, []):
                valid_subs.add(sub)
        subfunction_options = sorted(valid_subs)
    elif selected_teams:
        # If no specific function selected, show all subfunctions under the selected team(s)
        valid_subs = set()
        for team in selected_teams:
            # gather all subs under all functions of this team
            for func in team_to_functions.get(team, []):
                for sub in function_to_subs.get(func, []):
                    valid_subs.add(sub)
        subfunction_options = sorted(valid_subs)
    else:
        subfunction_options = subfunctions_list  # no filters => all subs
    selected_subfunctions = st.multiselect("Subfunction", options=subfunction_options, default=[])
    
    # Determine filtered result set
    # If no filter selected at all, include all companies
    if not selected_teams and not selected_functions and not selected_subfunctions:
        filtered_df = df_ds2.copy()
    else:
        # Apply filters (AND logic: entry must match all non-empty criteria)
        filtered_df = df_ds2.copy()
        if selected_teams:
            filtered_df = filtered_df[filtered_df["Team"].isin(selected_teams)]
        if selected_functions:
            filtered_df = filtered_df[filtered_df["Function"].isin(selected_functions)]
        if selected_subfunctions:
            filtered_df = filtered_df[filtered_df["Subfunction"].isin(selected_subfunctions)]
    # Aggregate results by company Name
    result_names = filtered_df["Name"].unique().tolist()
    result_count = len(result_names)
    st.caption(f"Found {result_count} matching company(s)")
    if result_count == 0:
        st.info("No companies match the selected criteria.")
    else:
        # Aggregate category labels per company (only from the filtered entries)
        agg = filtered_df.groupby("Name").agg(
            teams=('Team', lambda x: ", ".join(sorted({str(v).strip() for v in x.dropna() if str(v).strip() != ""}))),
            functions=('Function', lambda x: ", ".join(sorted({str(v).strip() for v in x.dropna() if str(v).strip() != ""}))),
            subfunctions=('Subfunction', lambda x: ", ".join(sorted({str(v).strip() for v in x.dropna() if str(v).strip() != ""})))
        ).reset_index()
        # Convert aggregated DataFrame to dictionary for quick lookup
        categories_by_name = {
            row["Name"]: {
                "teams": row["teams"],
                "functions": row["functions"],
                "subfunctions": row["subfunctions"]
            } for _, row in agg.iterrows()
        }

        # Interactive results table with category columns
        df_results = build_results_df(result_names, include_categories=True, categories_by_name=categories_by_name)
        render_results_df(df_results, show_categories=True)

elif search_mode == "Keyword":
    # Keyword/Description Search
    st.write("**Search company descriptions for keywords:**")
    col1, col2 = st.columns(2)
    term1 = col1.text_input("Keyword 1:").strip()
    term2 = col2.text_input("Keyword 2 (optional):").strip()
    # Only proceed if at least one term is provided
    if term1 or term2:
        # Filter companies whose description contains term1 or term2 (case-insensitive)
        terms = []
        if term1:
            terms.append(term1.lower())
        if term2:
            terms.append(term2.lower())
        result_names = []
        for name, info in details_by_name.items():
            desc = info.get("Description", "").lower()
            # Multiple terms are OR (match any term)
            if any(t in desc for t in terms):
                result_names.append(name)
        result_count = len(result_names)
        st.caption(f"Found {result_count} matching company(s)")
        if result_count == 0:
            st.info("No companies found for those keywords.")
        else:
            # Build and display results table (interactive)
            df_results = build_results_df(result_names, include_categories=False)
            render_results_df(df_results, show_categories=False)

